"""Convert per-day historical baseline XLSX/CSV data into a frontend JSON asset."""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import load_workbook

LOGGER = logging.getLogger("process_baseline_daily")

REPO_ROOT = Path(__file__).resolve().parents[4]
BACKEND_ROOT = REPO_ROOT / "CLISApp-backend"
FRONTEND_ROOT = REPO_ROOT / "CLISApp-frontend"
DATA_ROOT = BACKEND_ROOT / "data_pipeline" / "data"
RAW_BASELINE_DIR = DATA_ROOT / "raw" / "baseline"
DEFAULT_INPUT_PATH = RAW_BASELINE_DIR / "baseline_daily_indicators_by_doy_1890_1960.csv"
DEFAULT_YEARLY_INPUT_PATH = RAW_BASELINE_DIR / "baseline_ref_bmT_bmTmin_bmR_1890_1960.csv"
OUTPUT_PATH = FRONTEND_ROOT / "assets" / "data" / "climate_baseline_daily.json"

EXPECTED_SSC_COUNT = 2894
PERIOD = "1890-1960"
DESCRIPTION = "Per-day historical baseline (min/mean/max temp + rainfall) per SSC"
REQUIRED_COLUMNS = (
    "MERGED_ID",
    "baseline_mT",
    "baseline_MT",
    "baseline_rf",
    "baseline_tmean",
    "n_years",
)
DAY_COLUMN_ALIASES = ("day", "doy")
OPTIONAL_COLUMNS = ("suburb_name",)
SUPPORTED_INPUT_SUFFIXES = (".xlsx", ".csv")
FIELD_ORDER = ("tmean", "MT", "mT", "rf")
FIELD_UNITS = {"tmean": "°C", "MT": "°C", "mT": "°C", "rf": "mm/day"}
FIELD_RANGES = {
    "baseline_mT": (5.0, 30.0, "C"),
    "baseline_MT": (15.0, 45.0, "C"),
    "baseline_tmean": (10.0, 35.0, "C"),
    "baseline_rf": (0.0, 500.0, "mm"),
}
PARTIAL_DAY_THRESHOLD = 351
DAY_COUNT = 366


class BaselineDailyProcessingError(ValueError):
    """Raised when the daily baseline XLSX/CSV fails validation."""


@dataclass(frozen=True)
class DailyBaselineRecord:
    merged_id: int
    suburb_name: str
    day: int
    mT: float | None
    MT: float | None
    rf: float | None
    tmean: float | None
    n_years: int
    row_number: int

    @property
    def ssc_id(self) -> str:
        return f"ssc_{self.merged_id}"


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert per-day historical baseline XLSX or CSV into a static frontend JSON asset."
        ),
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help=(
            f"Daily baseline input path; .xlsx or .csv supported "
            f"(default: {DEFAULT_INPUT_PATH})"
        ),
    )
    parser.add_argument(
        "--yearly-input",
        type=Path,
        default=DEFAULT_YEARLY_INPUT_PATH,
        help=f"Yearly baseline CSV used for expected SSC coverage (default: {DEFAULT_YEARLY_INPUT_PATH})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT_PATH,
        help=f"JSON output path (default: {OUTPUT_PATH})",
    )
    return parser.parse_args(argv)


def _format_timestamp(value: datetime) -> str:
    normalized = value if value.tzinfo is not None else value.replace(tzinfo=UTC)
    return normalized.astimezone(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _default_generated_at(input_path: Path) -> datetime:
    return datetime.fromtimestamp(input_path.stat().st_mtime, UTC)


def _validate_required_columns(header_values: Sequence[Any]) -> dict[str, int]:
    fieldnames = [str(value).strip() if value is not None else "" for value in header_values]
    found_columns_by_name = {name: idx for idx, name in enumerate(fieldnames) if name}
    found_columns_label = ", ".join(name for name in fieldnames if name) or "<none>"

    missing_columns = [c for c in REQUIRED_COLUMNS if c not in found_columns_by_name]
    if missing_columns:
        raise BaselineDailyProcessingError(
            f"Missing required columns: {', '.join(missing_columns)}. "
            f"Found columns: {found_columns_label}"
        )

    day_alias = next(
        (name for name in DAY_COLUMN_ALIASES if name in found_columns_by_name), None
    )
    if day_alias is None:
        raise BaselineDailyProcessingError(
            f"Missing day column (expected one of: {', '.join(DAY_COLUMN_ALIASES)}). "
            f"Found columns: {found_columns_label}"
        )

    column_index: dict[str, int] = {c: found_columns_by_name[c] for c in REQUIRED_COLUMNS}
    column_index["day"] = found_columns_by_name[day_alias]
    for optional_column in OPTIONAL_COLUMNS:
        if optional_column in found_columns_by_name:
            column_index[optional_column] = found_columns_by_name[optional_column]
    return column_index


def _parse_int(raw_value: Any, *, column: str, row_number: int) -> int:
    if raw_value is None or raw_value == "":
        raise BaselineDailyProcessingError(f"Missing {column} value at row {row_number}")
    try:
        return int(raw_value)
    except (TypeError, ValueError) as exc:
        raise BaselineDailyProcessingError(
            f"Invalid {column} value '{raw_value}' at row {row_number}"
        ) from exc


def _parse_float(
    raw_value: Any,
    *,
    column: str,
    row_number: int,
    merged_id: int,
    day: int,
) -> float | None:
    if raw_value is None or raw_value == "":
        raise BaselineDailyProcessingError(
            f"Missing {column} value at row {row_number} (MERGED_ID={merged_id})"
        )
    if isinstance(raw_value, str) and raw_value.strip().upper() == "NA":
        LOGGER.warning(
            "NA %s value at row %d (MERGED_ID=%s, day=%d); writing null for this day",
            column,
            row_number,
            merged_id,
            day,
        )
        return None
    try:
        value = float(raw_value)
    except (TypeError, ValueError) as exc:
        raise BaselineDailyProcessingError(
            f"Invalid {column} value '{raw_value}' at row {row_number} (MERGED_ID={merged_id})"
        ) from exc

    warning_min, warning_max, unit = FIELD_RANGES[column]
    if not warning_min <= value <= warning_max:
        LOGGER.warning(
            "%s value %.2f at row %d (MERGED_ID=%s) outside expected range %.0f-%.0f %s",
            column,
            value,
            row_number,
            merged_id,
            warning_min,
            warning_max,
            unit,
        )

    return value


def _parse_n_years(raw_value: Any, *, row_number: int, merged_id: int) -> int:
    if raw_value is None or raw_value == "":
        raise BaselineDailyProcessingError(
            f"Missing n_years value at row {row_number} (MERGED_ID={merged_id})"
        )
    try:
        return int(raw_value)
    except (TypeError, ValueError) as exc:
        raise BaselineDailyProcessingError(
            f"Invalid n_years value '{raw_value}' at row {row_number} (MERGED_ID={merged_id})"
        ) from exc


def _records_from_rows(
    rows: Iterable[Sequence[Any]],
    column_index: dict[str, int],
) -> tuple[DailyBaselineRecord, ...]:
    records: list[DailyBaselineRecord] = []
    seen_day_by_id: dict[tuple[int, int], int] = {}
    suburb_idx = column_index.get("suburb_name")

    for row_number, row in enumerate(rows, start=2):
        merged_id = _parse_int(
            row[column_index["MERGED_ID"]],
            column="MERGED_ID",
            row_number=row_number,
        )
        day = _parse_int(row[column_index["day"]], column="day", row_number=row_number)
        if not 1 <= day <= DAY_COUNT:
            raise BaselineDailyProcessingError(
                f"Invalid day value {day} at row {row_number} (MERGED_ID={merged_id}); "
                "expected 1-366"
            )

        seen_key = (merged_id, day)
        if seen_key in seen_day_by_id:
            raise BaselineDailyProcessingError(
                f"Duplicate MERGED_ID/day {merged_id}/{day} found at rows "
                f"{seen_day_by_id[seen_key]} and {row_number}"
            )
        seen_day_by_id[seen_key] = row_number

        suburb_name = (
            str(row[suburb_idx] or "").strip() if suburb_idx is not None else ""
        )

        records.append(
            DailyBaselineRecord(
                merged_id=merged_id,
                suburb_name=suburb_name,
                day=day,
                mT=_parse_float(
                    row[column_index["baseline_mT"]],
                    column="baseline_mT",
                    row_number=row_number,
                    merged_id=merged_id,
                    day=day,
                ),
                MT=_parse_float(
                    row[column_index["baseline_MT"]],
                    column="baseline_MT",
                    row_number=row_number,
                    merged_id=merged_id,
                    day=day,
                ),
                rf=_parse_float(
                    row[column_index["baseline_rf"]],
                    column="baseline_rf",
                    row_number=row_number,
                    merged_id=merged_id,
                    day=day,
                ),
                tmean=_parse_float(
                    row[column_index["baseline_tmean"]],
                    column="baseline_tmean",
                    row_number=row_number,
                    merged_id=merged_id,
                    day=day,
                ),
                n_years=_parse_n_years(
                    row[column_index["n_years"]],
                    row_number=row_number,
                    merged_id=merged_id,
                ),
                row_number=row_number,
            )
        )

    return tuple(records)


def _load_xlsx_records(input_path: Path) -> tuple[DailyBaselineRecord, ...]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["in"]:
            raise BaselineDailyProcessingError(
                f"Expected a single sheet named 'in'. Found: {', '.join(workbook.sheetnames)}"
            )

        sheet = workbook["in"]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration as exc:
            raise BaselineDailyProcessingError(
                f"XLSX is missing a header row. Required columns: {', '.join(REQUIRED_COLUMNS)}"
            ) from exc

        column_index = _validate_required_columns(header_values)
        return _records_from_rows(rows, column_index)
    finally:
        workbook.close()


def _load_csv_records(input_path: Path) -> tuple[DailyBaselineRecord, ...]:
    with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            header_values = next(reader)
        except StopIteration as exc:
            raise BaselineDailyProcessingError(
                f"CSV is missing a header row. Required columns: {', '.join(REQUIRED_COLUMNS)}"
            ) from exc
        column_index = _validate_required_columns(header_values)
        return _records_from_rows(reader, column_index)


def load_daily_baseline_records(input_path: Path) -> tuple[DailyBaselineRecord, ...]:
    if not input_path.exists():
        raise FileNotFoundError(f"Daily baseline file not found: {input_path}")

    suffix = input_path.suffix.lower()
    if suffix == ".xlsx":
        return _load_xlsx_records(input_path)
    if suffix == ".csv":
        return _load_csv_records(input_path)
    raise BaselineDailyProcessingError(
        f"Unsupported input file extension '{suffix}' for {input_path.name}. "
        f"Expected one of: {', '.join(SUPPORTED_INPUT_SUFFIXES)}."
    )


def load_expected_merged_ids(yearly_input_path: Path) -> set[int]:
    if not yearly_input_path.exists():
        raise FileNotFoundError(f"Yearly baseline CSV not found: {yearly_input_path}")

    expected_ids: set[int] = set()
    with yearly_input_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None or "MERGED_ID" not in reader.fieldnames:
            raise BaselineDailyProcessingError("Yearly baseline CSV is missing MERGED_ID column")
        for row_number, row in enumerate(reader, start=2):
            raw_merged_id = (row.get("MERGED_ID") or "").strip()
            try:
                expected_ids.add(int(raw_merged_id))
            except ValueError as exc:
                raise BaselineDailyProcessingError(
                    f"Invalid MERGED_ID '{raw_merged_id}' in yearly CSV at row {row_number}"
                ) from exc
    return expected_ids


def _group_records_by_merged_id(
    records: Iterable[DailyBaselineRecord],
) -> dict[int, list[DailyBaselineRecord]]:
    grouped: dict[int, list[DailyBaselineRecord]] = {}
    for record in records:
        grouped.setdefault(record.merged_id, []).append(record)
    return grouped


def _build_baseline_arrays(records: Sequence[DailyBaselineRecord]) -> dict[str, list[float | None]]:
    values: dict[str, list[float | None]] = {
        field: [None] * DAY_COUNT for field in FIELD_ORDER
    }
    for record in records:
        day_index = record.day - 1
        values["tmean"][day_index] = round(record.tmean, 1) if record.tmean is not None else None
        values["MT"][day_index] = round(record.MT, 1) if record.MT is not None else None
        values["mT"][day_index] = round(record.mT, 1) if record.mT is not None else None
        values["rf"][day_index] = round(record.rf, 1) if record.rf is not None else None
    return values


def _truncation_note(
    missing_ssc_ids: Sequence[int],
    partial_ssc_ids: dict[str, int],
    *,
    source_name: str,
) -> str:
    missing_count = len(missing_ssc_ids)
    partial_count = len(partial_ssc_ids)
    is_xlsx = source_name.lower().endswith(".xlsx")
    if is_xlsx and missing_count == 29 and partial_ssc_ids == {"33228": 351}:
        return (
            "Source xlsx hit Excel's 1,048,576-row limit; 29 SSCs missing, 1 partial. "
            "Flagged to Javier 2026-05-08; see DEC-11."
        )
    source_label = "Source xlsx may be truncated" if is_xlsx else "Source data may be incomplete"
    if missing_count or partial_count:
        return f"{source_label}; {missing_count} SSCs missing, {partial_count} partial."
    if is_xlsx:
        return "Source xlsx may be truncated; no missing SSCs detected against yearly baseline."
    return "Daily baseline complete; all SSCs covered against yearly baseline."


def build_payload(
    records: Sequence[DailyBaselineRecord],
    *,
    expected_merged_ids: set[int],
    source_name: str,
    generated_at: datetime | None = None,
    expected_ssc_count: int = EXPECTED_SSC_COUNT,
) -> dict[str, Any]:
    timestamp = _format_timestamp(generated_at or datetime.now(UTC))
    grouped_records = _group_records_by_merged_id(records)
    na_value_count = sum(
        value is None
        for record in records
        for value in (record.tmean, record.MT, record.mT, record.rf)
    )

    partial_ssc_ids: dict[str, int] = {}
    baselines: dict[str, dict[str, list[float | None]]] = {}
    for merged_id in sorted(grouped_records):
        region_records = grouped_records[merged_id]
        present_days = {record.day for record in region_records}
        day_count = len(present_days)
        if day_count < PARTIAL_DAY_THRESHOLD:
            raise BaselineDailyProcessingError(
                f"MERGED_ID {merged_id} has {day_count} days; expected at least {PARTIAL_DAY_THRESHOLD}"
            )
        if day_count < DAY_COUNT:
            partial_ssc_ids[str(merged_id)] = day_count
        baselines[f"ssc_{merged_id}"] = _build_baseline_arrays(region_records)

    present_merged_ids = set(grouped_records)
    missing_ssc_ids = sorted(expected_merged_ids - present_merged_ids)
    metadata = {
        "source": source_name,
        "period": PERIOD,
        "description": DESCRIPTION,
        "generated": timestamp,
        "ssc_count": len(present_merged_ids),
        "expected_ssc_count": expected_ssc_count,
        "missing_ssc_ids": missing_ssc_ids,
        "partial_ssc_ids": partial_ssc_ids,
        "truncation_note": _truncation_note(
            missing_ssc_ids, partial_ssc_ids, source_name=source_name
        ),
        "field_units": FIELD_UNITS,
    }
    if na_value_count:
        metadata["na_value_count"] = na_value_count
    return {
        "metadata": metadata,
        "baselines": baselines,
    }


def write_json_atomic(output_path: Path, payload: dict[str, Any]) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    json_text = json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n"
    temp_path: Path | None = None

    try:
        with NamedTemporaryFile(
            "w",
            encoding="utf-8",
            dir=output_path.parent,
            prefix=f".{output_path.stem}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            temp_path = Path(handle.name)
            handle.write(json_text)
            handle.flush()
            os.fsync(handle.fileno())

        temp_path.replace(output_path)

        with output_path.open("r", encoding="utf-8") as handle:
            json.load(handle)

        return output_path.stat().st_size
    finally:
        if temp_path is not None and temp_path.exists():
            temp_path.unlink(missing_ok=True)


def _print_truncation_report(payload: dict[str, Any], *, data_row_count: int) -> None:
    metadata = payload["metadata"]
    expected_rows = metadata["expected_ssc_count"] * DAY_COUNT
    row_shortfall = expected_rows - data_row_count
    print("Truncation report:")
    print(f"  source_rows: {data_row_count}")
    print(f"  expected_rows: {expected_rows}")
    print(f"  row_shortfall: {row_shortfall}")
    print(f"  ssc_count: {metadata['ssc_count']} / {metadata['expected_ssc_count']}")
    print(f"  na_value_count: {metadata.get('na_value_count', 0)}")
    print(f"  missing_ssc_ids: {metadata['missing_ssc_ids']}")
    print(f"  partial_ssc_ids: {metadata['partial_ssc_ids']}")
    print(f"  note: {metadata['truncation_note']}")


def process_baseline_daily(
    *,
    input_path: Path = DEFAULT_INPUT_PATH,
    yearly_input_path: Path = DEFAULT_YEARLY_INPUT_PATH,
    output_path: Path = OUTPUT_PATH,
    generated_at: datetime | None = None,
    expected_ssc_count: int = EXPECTED_SSC_COUNT,
) -> dict[str, Any]:
    records = load_daily_baseline_records(input_path)
    expected_merged_ids = load_expected_merged_ids(yearly_input_path)
    if len(expected_merged_ids) != expected_ssc_count:
        LOGGER.warning(
            "Expected %d yearly SSC ids, got %d. Continuing.",
            expected_ssc_count,
            len(expected_merged_ids),
        )

    payload = build_payload(
        records,
        expected_merged_ids=expected_merged_ids,
        source_name=input_path.name,
        generated_at=generated_at or _default_generated_at(input_path),
        expected_ssc_count=expected_ssc_count,
    )
    _print_truncation_report(payload, data_row_count=len(records))
    file_size_bytes = write_json_atomic(output_path, payload)

    LOGGER.info("Processed %d daily baseline regions", payload["metadata"]["ssc_count"])
    LOGGER.info("Daily baseline JSON written to %s (%d bytes)", output_path, file_size_bytes)

    return {
        "ssc_count": payload["metadata"]["ssc_count"],
        "output_path": output_path,
        "file_size_bytes": file_size_bytes,
        "payload": payload,
    }


def main(argv: Sequence[str] | None = None) -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    args = parse_args(argv)

    try:
        result = process_baseline_daily(
            input_path=args.input,
            yearly_input_path=args.yearly_input,
            output_path=args.output,
        )
    except (BaselineDailyProcessingError, FileNotFoundError) as exc:
        LOGGER.error("%s", exc)
        raise SystemExit(1) from exc

    print(f"Total regions: {result['ssc_count']}")
    print(f"Output path: {result['output_path']}")
    print(f"File size: {result['file_size_bytes']} bytes")


if __name__ == "__main__":
    main()
